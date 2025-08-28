import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Create test data for VAPT dashboard
def create_test_excel():
    # Create a new workbook
    wb = Workbook()
    
    # Remove default worksheet
    wb.remove(wb.active)
    
    # Create Proposal sheet
    proposal_data = {
        'S.No': [1, 2, 3],
        'Domain': ['Web Application', 'Network Infrastructure', 'Mobile Application'],
        'Different Avenues of the testing scope': [
            'Authentication, Authorization, Input Validation',
            'Firewall, IDS/IPS, Network Segmentation',
            'API Security, Data Storage, Communication'
        ],
        'Overview of the methodology': [
            'OWASP Top 10 based testing',
            'NIST Cybersecurity Framework',
            'OWASP Mobile Top 10'
        ],
        'Key Dilevirables': [
            'Vulnerability Report, Remediation Guide',
            'Network Security Assessment',
            'Mobile Security Report'
        ],
        'Key Dependencies': [
            'Test Environment Access',
            'Network Documentation',
            'Source Code Access'
        ],
        'Est.testdate': ['2024-03-15', '2024-04-20', '2024-05-10'],
        'Est.startdate': ['2024-03-01', '2024-04-01', '2024-05-01'],
        'Timeline of testing': ['2 weeks', '3 weeks', '2 weeks'],
        'Timeline of remediation': ['1 week', '2 weeks', '1 week'],
        'Stake holder': ['IT Team', 'Network Team', 'Development Team'],
        'Last tested year': ['2024-25', '2024-25', '2024-25'],
        'Requirement statements': [
            'Comprehensive web app security testing',
            'Network infrastructure assessment',
            'Mobile application security review'
        ]
    }
    
    # Create Scope sheet
    scope_data = {
        'S.No': [1, 2, 3],
        'Penetration testing': ['Web App Penetration Testing', 'Network Penetration Testing', 'Mobile App Testing'],
        'Description': [
            'Testing web application for security vulnerabilities',
            'Testing network infrastructure for security gaps',
            'Testing mobile application for security issues'
        ],
        'Key Dilevirables': [
            'Detailed vulnerability report',
            'Network security assessment report',
            'Mobile security test report'
        ],
        'Key Dependencies': [
            'Application access and credentials',
            'Network topology and access',
            'Mobile app binary and source code'
        ],
        'Est.testdate': ['2024-03-15', '2024-04-20', '2024-05-10'],
        'Est.startdate': ['2024-03-01', '2024-04-01', '2024-05-01'],
        'Timeline of testing': ['2 weeks', '3 weeks', '2 weeks'],
        'Timeline of remediation': ['1 week', '2 weeks', '1 week'],
        'Stake holder': ['IT Team', 'Network Team', 'Development Team'],
        'Last tested year': ['2024-25', '2024-25', '2024-25'],
        'Requirement statements': [
            'OWASP Top 10 compliance testing',
            'PCI DSS network compliance',
            'OWASP Mobile Top 10 testing'
        ],
        'Impact of these tests': ['HIGH', 'CRITICAL', 'MEDIUM']
    }
    
    # Create VAPT Results sheet
    vapt_results_data = {
        'Vulnerabilities ID': ['VAPT-2024-001', 'VAPT-2024-002', 'VAPT-2024-003'],
        'Name of vulnerability': [
            'SQL Injection in Login Form',
            'Unencrypted Network Traffic',
            'Insecure Data Storage'
        ],
        'Description': [
            'SQL injection vulnerability found in user login form allowing data extraction',
            'Network traffic between client and server is not encrypted',
            'Sensitive data stored in plain text on mobile device'
        ],
        'Tested environment': ['PRODUCTION', 'STAGING', 'DEVELOPMENT'],
        'Result': ['FAIL', 'FAIL', 'FAIL'],
        'Reporting evidence': [
            'Screenshot showing successful SQL injection',
            'Network capture showing unencrypted traffic',
            'File system analysis showing plain text data'
        ],
        'Source of identifcation': [
            'Manual Testing',
            'Network Analysis',
            'Code Review'
        ],
        'Criticiality based on CVSS score': ['HIGH', 'MEDIUM', 'HIGH'],
        'Criticiality based on business context': ['CRITICAL', 'HIGH', 'MEDIUM'],
        'Stake holders': ['IT Team, Security Team', 'Network Team', 'Development Team'],
        'Remarks from stakeholders': [
            'Critical fix required immediately',
            'Implement SSL/TLS encryption',
            'Update data encryption methods'
        ]
    }
    
    # Add sheets to workbook
    ws1 = wb.create_sheet("Proposal (25-26)")
    df1 = pd.DataFrame(proposal_data)
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws1.append(r)
    
    ws2 = wb.create_sheet("Scope (24-25)")
    df2 = pd.DataFrame(scope_data)
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws2.append(r)
    
    ws3 = wb.create_sheet("VAPT Results")
    df3 = pd.DataFrame(vapt_results_data)
    for r in dataframe_to_rows(df3, index=False, header=True):
        ws3.append(r)
    
    # Save the file
    wb.save('test_vapt_data.xlsx')
    print("Test Excel file created: test_vapt_data.xlsx")

if __name__ == "__main__":
    create_test_excel()